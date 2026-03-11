#!/usr/bin/env python3
"""
Export Soya Oil, Sunflower Oil, and Palm Oil Material Consumption Report to Excel
Month-by-month breakdown from Jan 1, 2025 to Jan 1, 2026
Creates Excel file with monthly data for each material
Run from NFM-backend directory: python export_oils_monthly_to_excel.py
"""

import sys
import os
from datetime import datetime, timedelta
from sqlalchemy import func, or_, extract
from flask import Flask
import pandas as pd
from collections import defaultdict

# Add the backend directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import Config
from extensions import db
from models.kpi_material import KPIMaterial

def export_oils_monthly_to_excel():
    """Export Soya Oil, Sunflower Oil, and Palm Oil data to Excel with month-by-month breakdown"""
    
    # Create a minimal Flask app for database access
    app = Flask(__name__)
    app.config.from_object(Config)
    db.init_app(app)
    
    with app.app_context():
        # Date range: Jan 1, 2025 7:00 AM to Jan 1, 2026 7:00 AM
        start_date = datetime(2025, 1, 1, 7, 0, 0)
        end_date = datetime(2026, 1, 1, 7, 0, 0)
        
        # Materials to query
        materials_to_query = ['Soya Oil', 'FAT', 'Palm Oil']
        
        # First, check what materials actually exist
        print("Checking which materials exist in database...")
        existing_materials = set()
        all_materials_query = KPIMaterial.query.filter(
            KPIMaterial.batch_act_start >= start_date,
            KPIMaterial.batch_act_start <= end_date,
            func.lower(KPIMaterial.product_name) != 'not selected'
        )
        
        for record in all_materials_query.all():
            mat_name = record.material_name or ''
            # Check if it matches any of our query materials (case-insensitive)
            for query_mat in materials_to_query:
                if query_mat.lower() in mat_name.lower():
                    existing_materials.add(query_mat)
                    break
        
        print(f"Materials found in database: {sorted(existing_materials)}")
        missing_materials = set(materials_to_query) - existing_materials
        if missing_materials:
            print(f"WARNING: These materials not found in database: {sorted(missing_materials)}")
            print("They will show as 0.00 in the report.")
        
        print("=" * 80)
        print("EXPORTING OILS MATERIAL CONSUMPTION REPORT (MONTHLY) TO EXCEL")
        print("=" * 80)
        print(f"Date Range: {start_date} to {end_date}")
        print(f"Materials: {', '.join(materials_to_query)}")
        print("=" * 80)
        print()
        
        try:
            # Query for all specified oil materials
            print("Querying database for oil materials data...")
            materials_query = KPIMaterial.query.filter(
                KPIMaterial.batch_act_start >= start_date,
                KPIMaterial.batch_act_start <= end_date,
                or_(
                    func.lower(KPIMaterial.material_name).like('%soya oil%'),
                    func.lower(KPIMaterial.material_name).like('%fat%'),
                    func.lower(KPIMaterial.material_name).like('%palm oil%')
                ),
                func.lower(KPIMaterial.product_name) != 'not selected'
            )
            
            all_records = materials_query.all()
            print(f"Found {len(all_records)} total records")
            
            if len(all_records) == 0:
                print("[X] No records found. Exiting.")
                return 1
            
            # Group data by material name and month
            print("Aggregating data by material and month...")
            monthly_data = defaultdict(lambda: defaultdict(lambda: {'planned': 0.0, 'actual': 0.0}))
            
            for record in all_records:
                material_name = record.material_name or 'Unknown'
                batch_date = record.batch_act_start
                
                if batch_date:
                    # Get year and month
                    year_month = batch_date.strftime('%Y-%m')
                    month_name = batch_date.strftime('%B %Y')
                    
                    # Normalize material name (handle case variations)
                    normalized_name = None
                    for mat in materials_to_query:
                        if mat.lower() in material_name.lower():
                            normalized_name = mat
                            break
                    
                    if normalized_name:
                        planned = float(record.setpoint_float or 0)
                        actual = float(record.actual_value_float or 0)
                        
                        monthly_data[normalized_name][month_name]['planned'] += planned
                        monthly_data[normalized_name][month_name]['actual'] += actual
            
            # Generate all months in the range
            all_months = []
            current_date = start_date.replace(day=1, hour=0, minute=0, second=0)
            while current_date < end_date:
                month_name = current_date.strftime('%B %Y')
                all_months.append(month_name)
                # Move to next month
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)
            
            # Create report data structure - organized by month first, then materials
            report_data = []
            
            # For each month, show all materials
            for month in all_months:
                # Add month header row (will span all columns)
                report_data.append({
                    'Material Name': month,
                    'Planned (KG)': '',
                    'Actual (KG)': '',
                    'Difference %': ''
                })
                
                # Add all materials for this month
                month_total_planned = 0.0
                month_total_actual = 0.0
                
                for material in materials_to_query:
                    if month in monthly_data[material]:
                        data = monthly_data[material][month]
                        planned = data['planned']
                        actual = data['actual']
                        month_total_planned += planned
                        month_total_actual += actual
                        
                        # Calculate difference percentage
                        if planned != 0:
                            difference_percent = abs(((actual - planned) / planned) * 100)
                        else:
                            difference_percent = 0.0
                        
                        report_data.append({
                            'Material Name': material,
                            'Planned (KG)': round(planned, 2),
                            'Actual (KG)': round(actual, 2),
                            'Difference %': round(difference_percent, 2)
                        })
                    else:
                        # Add row with zeros if no data for this month
                        report_data.append({
                            'Material Name': material,
                            'Planned (KG)': 0.0,
                            'Actual (KG)': 0.0,
                            'Difference %': 0.0
                        })
                
                # Add month total row
                month_total_diff = abs(((month_total_actual - month_total_planned) / month_total_planned * 100)) if month_total_planned != 0 else 0.0
                report_data.append({
                    'Material Name': f"{month} - TOTAL",
                    'Planned (KG)': round(month_total_planned, 2),
                    'Actual (KG)': round(month_total_actual, 2),
                    'Difference %': round(month_total_diff, 2)
                })
                
                # Add empty row for spacing
                report_data.append({
                    'Material Name': '',
                    'Planned (KG)': '',
                    'Actual (KG)': '',
                    'Difference %': ''
                })
            
            # Add overall summary at the end
            report_data.append({
                'Material Name': '=== OVERALL SUMMARY ===',
                'Planned (KG)': '',
                'Actual (KG)': '',
                'Difference %': ''
            })
            
            for material in materials_to_query:
                material_total_planned = sum(d['planned'] for d in monthly_data[material].values())
                material_total_actual = sum(d['actual'] for d in monthly_data[material].values())
                material_total_diff = abs(((material_total_actual - material_total_planned) / material_total_planned * 100)) if material_total_planned != 0 else 0.0
                
                report_data.append({
                    'Material Name': f"{material} - TOTAL",
                    'Planned (KG)': round(material_total_planned, 2),
                    'Actual (KG)': round(material_total_actual, 2),
                    'Difference %': round(material_total_diff, 2)
                })
            
            # Create DataFrame
            df = pd.DataFrame(report_data)
            
            # Create Excel file
            output_filename = f"Oils_Monthly_Material_Consumption_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_filename)
            
            print(f"\nCreating Excel file: {output_filename}")
            
            # Write to Excel with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Monthly Consumption', index=False)
                
                # Get the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Monthly Consumption']
                
                # Format header row
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Format data rows
                data_font = Font(size=10)
                number_alignment = Alignment(horizontal="right", vertical="center")
                text_alignment = Alignment(horizontal="left", vertical="center")
                total_font = Font(size=10, bold=True)
                total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                current_row = 2
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    cell_value = str(row[0].value) if row[0].value else ''
                    is_total_row = 'TOTAL' in cell_value
                    is_month_header = '===' in cell_value
                    is_summary_header = 'OVERALL SUMMARY' in cell_value
                    is_empty_row = not cell_value or cell_value == ''
                    
                    for cell in row:
                        if not is_empty_row:
                            cell.border = thin_border
                        
                        if is_month_header or is_summary_header:
                            cell.font = Font(bold=True, size=11, color="FFFFFF")
                            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif is_total_row:
                            cell.font = total_font
                            cell.fill = total_fill
                            if cell.column in [3, 4, 5]:  # Planned, Actual, Difference columns
                                cell.alignment = number_alignment
                            else:
                                cell.alignment = text_alignment
                        elif not is_empty_row:
                            if cell.column in [3, 4, 5]:  # Planned, Actual, Difference columns
                                cell.alignment = number_alignment
                                cell.font = data_font
                            else:  # Material Name and Month columns
                                cell.alignment = text_alignment
                                cell.font = data_font
                    
                    current_row += 1
                
                # Auto-adjust column widths
                column_widths = {
                    'A': 30,  # Material Name
                    'B': 15,  # Planned (KG)
                    'C': 15,  # Actual (KG)
                    'D': 15   # Difference %
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Add title row
                worksheet.insert_rows(1)
                worksheet.merge_cells('A1:D1')
                title_cell = worksheet['A1']
                title_cell.value = f"Material Consumption Report - Monthly Breakdown\nMaterials: {', '.join(materials_to_query)}\nDate Range: {start_date.strftime('%Y-%m-%d %H:%M')} to {end_date.strftime('%Y-%m-%d %H:%M')}"
                title_cell.font = Font(bold=True, size=12)
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                
                # Merge month header cells and format them
                current_row = 2  # Start after title and column headers
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    cell_value = str(row[0].value) if row[0].value else ''
                    # Check if this is a month header (contains month name pattern like "January 2025" but not "TOTAL")
                    is_month_header = (cell_value and 
                                     'TOTAL' not in cell_value and 
                                     'OVERALL SUMMARY' not in cell_value and
                                     cell_value not in materials_to_query and
                                     cell_value != '' and
                                     any(month in cell_value for month in all_months))
                    
                    if is_month_header:
                        # Merge cells for month header (A to D columns)
                        worksheet.merge_cells(f'A{current_row}:D{current_row}')
                        # Format only the first cell (merged)
                        cell = row[0]
                        cell.border = thin_border
                        cell.font = Font(bold=True, size=11, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    current_row += 1
                
                # Adjust header row position
                worksheet.row_dimensions[2].height = 25
                worksheet.row_dimensions[1].height = 50
            
            print(f"\n[OK] Excel file created successfully!")
            print(f"File location: {output_path}")
            print(f"\nSummary:")
            
            # Print summary for each material
            for material in materials_to_query:
                if material in monthly_data:
                    total_planned = sum(d['planned'] for d in monthly_data[material].values())
                    total_actual = sum(d['actual'] for d in monthly_data[material].values())
                    months_with_data = len([m for m in all_months if m in monthly_data[material]])
                    print(f"\n  {material}:")
                    print(f"    Total Planned: {total_planned:.2f} KG")
                    print(f"    Total Actual: {total_actual:.2f} KG")
                    print(f"    Months with data: {months_with_data} out of {len(all_months)}")
            
            print("=" * 80)
            return 0
            
        except Exception as e:
            print(f"\n[ERROR] {str(e)}")
            import traceback
            traceback.print_exc()
            return 1

if __name__ == "__main__":
    exit_code = export_oils_monthly_to_excel()
    sys.exit(exit_code)

