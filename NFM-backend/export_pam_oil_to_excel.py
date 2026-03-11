#!/usr/bin/env python3
"""
Export PAM Oil and PAM Oil Ton Material Consumption Report to Excel
Creates Excel file with Material Name, Code, Planned (KG), Actual (KG), Difference %
Run from NFM-backend directory: python export_pam_oil_to_excel.py
"""

import sys
import os
from datetime import datetime
from sqlalchemy import func, or_
from flask import Flask
import pandas as pd

# Add the backend directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import Config
from extensions import db
from models.kpi_material import KPIMaterial

def export_pam_oil_to_excel():
    """Export PAM Oil and PAM Oil Ton data to Excel in Material Consumption Report format"""
    
    # Create a minimal Flask app for database access
    app = Flask(__name__)
    app.config.from_object(Config)
    db.init_app(app)
    
    with app.app_context():
        # Date range: Jan 1, 2025 7:00 AM to Jan 1, 2026 7:00 AM
        start_date = datetime(2025, 1, 1, 7, 0, 0)
        end_date = datetime(2026, 1, 1, 7, 0, 0)
        
        print("=" * 80)
        print("EXPORTING PAM OIL MATERIAL CONSUMPTION REPORT TO EXCEL")
        print("=" * 80)
        print(f"Date Range: {start_date} to {end_date}")
        print("=" * 80)
        print()
        
        try:
            # Query for all PAM/Palm Oil related materials
            print("Querying database for PAM Oil and PAM Oil Ton data...")
            materials_query = KPIMaterial.query.filter(
                KPIMaterial.batch_act_start >= start_date,
                KPIMaterial.batch_act_start <= end_date,
                or_(
                    func.lower(KPIMaterial.material_name).like('%pam oil%'),
                    func.lower(KPIMaterial.material_name).like('%palm oil%')
                ),
                func.lower(KPIMaterial.product_name) != 'not selected'
            )
            
            all_records = materials_query.all()
            print(f"Found {len(all_records)} total records")
            
            if len(all_records) == 0:
                print("[X] No records found. Exiting.")
                return 1
            
            # Aggregate by material (same logic as Material Consumption Report)
            print("Aggregating data by material...")
            material_groups = {}
            
            for record in all_records:
                material_name = record.material_name or 'Unknown'
                
                if material_name not in material_groups:
                    material_groups[material_name] = {
                        'material_name': material_name,
                        'code': record.material_code or '',
                        'planned_kg': 0.0,
                        'actual_kg': 0.0
                    }
                
                # Sum up planned and actual values
                material_groups[material_name]['planned_kg'] += float(record.setpoint_float or 0)
                material_groups[material_name]['actual_kg'] += float(record.actual_value_float or 0)
            
            # Calculate difference percentage for each material
            report_data = []
            for material_name, data in material_groups.items():
                planned = data['planned_kg']
                actual = data['actual_kg']
                
                # Calculate difference percentage (absolute value)
                if planned != 0:
                    difference_percent = abs(((actual - planned) / planned) * 100)
                else:
                    difference_percent = 0.0
                
                report_data.append({
                    'Material Name': data['material_name'],
                    'Code': data['code'],
                    'Planned (KG)': round(planned, 2),
                    'Actual (KG)': round(actual, 2),
                    'Difference %': round(difference_percent, 2)
                })
            
            # Sort by material name
            report_data.sort(key=lambda x: x['Material Name'])
            
            # Create DataFrame
            df = pd.DataFrame(report_data)
            
            # Create Excel file
            output_filename = f"PAM_Oil_Material_Consumption_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_filename)
            
            print(f"\nCreating Excel file: {output_filename}")
            
            # Write to Excel with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Material Consumption', index=False)
                
                # Get the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Material Consumption']
                
                # Format header row
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Format data rows
                data_font = Font(size=10)
                number_alignment = Alignment(horizontal="right", vertical="center")
                text_alignment = Alignment(horizontal="left", vertical="center")
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.border = thin_border
                        if cell.column in [3, 4, 5]:  # Planned, Actual, Difference columns
                            cell.alignment = number_alignment
                            cell.font = data_font
                        else:  # Material Name and Code columns
                            cell.alignment = text_alignment
                            cell.font = data_font
                
                # Auto-adjust column widths
                column_widths = {
                    'A': 30,  # Material Name
                    'B': 15,  # Code
                    'C': 15,  # Planned (KG)
                    'D': 15,  # Actual (KG)
                    'E': 15   # Difference %
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Add title row
                worksheet.insert_rows(1)
                worksheet.merge_cells('A1:E1')
                title_cell = worksheet['A1']
                title_cell.value = f"Material Consumption Report - PAM Oil & PAM Oil Ton\nDate Range: {start_date.strftime('%Y-%m-%d %H:%M')} to {end_date.strftime('%Y-%m-%d %H:%M')}"
                title_cell.font = Font(bold=True, size=12)
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                
                # Adjust header row position
                worksheet.row_dimensions[2].height = 25
            
            print(f"\n[OK] Excel file created successfully!")
            print(f"File location: {output_path}")
            print(f"\nSummary:")
            print(f"  Total materials: {len(report_data)}")
            print(f"  Total Planned: {sum(r['Planned (KG)'] for r in report_data):.2f} KG")
            print(f"  Total Actual: {sum(r['Actual (KG)'] for r in report_data):.2f} KG")
            
            # Print material breakdown
            print(f"\nMaterial Breakdown:")
            for item in report_data:
                print(f"  - {item['Material Name']}: {item['Planned (KG)']:.2f} KG planned, {item['Actual (KG)']:.2f} KG actual, {item['Difference %']:.2f}% difference")
            
            print("=" * 80)
            return 0
            
        except Exception as e:
            print(f"\n[ERROR] {str(e)}")
            import traceback
            traceback.print_exc()
            return 1

if __name__ == "__main__":
    exit_code = export_pam_oil_to_excel()
    sys.exit(exit_code)


